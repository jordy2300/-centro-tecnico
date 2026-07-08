import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET
from django.http import JsonResponse, HttpResponse
from io import BytesIO
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone

from .models import Material, Cuadrilla, Solicitud, ItemSolicitud
from .services import importar_materiales, exportar_solicitudes_excel


# ── PANEL ADMIN ──────────────────────────────────────────────────────────────

@login_required
def panel(request):
    total = Solicitud.objects.count()
    pendientes = Solicitud.objects.filter(estado='pendiente').count()
    aprobadas = Solicitud.objects.filter(estado='aprobada').count()
    en_entrega = Solicitud.objects.filter(estado='en_entrega').count()
    recientes = Solicitud.objects.select_related('cuadrilla').all()[:10]
    return render(request, 'materiales/panel.html', {
        'total': total, 'pendientes': pendientes,
        'aprobadas': aprobadas, 'en_entrega': en_entrega,
        'recientes': recientes,
    })


@login_required
def lista_solicitudes(request):
    qs = Solicitud.objects.select_related('cuadrilla').prefetch_related('items__material')
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    movil = request.GET.get('movil', '')

    if q:
        qs = qs.filter(Q(cuadrilla__nombre__icontains=q) | Q(cuadrilla__movil__icontains=q))
    if estado:
        qs = qs.filter(estado=estado)
    if desde:
        qs = qs.filter(fecha_solicitud__gte=desde)
    if hasta:
        qs = qs.filter(fecha_solicitud__lte=hasta)
    if movil:
        qs = qs.filter(cuadrilla__movil__icontains=movil)

    cuadrillas = Cuadrilla.objects.filter(activo=True)
    return render(request, 'materiales/lista_solicitudes.html', {
        'solicitudes': qs[:300], 'total': qs.count(),
        'cuadrillas': cuadrillas,
        'filtros': {'q': q, 'estado': estado, 'desde': desde, 'hasta': hasta, 'movil': movil},
        'estados': Solicitud.ESTADO_CHOICES,
    })


@login_required
def detalle_solicitud(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    return render(request, 'materiales/detalle_solicitud.html', {'sol': sol})


@login_required
@require_POST
def aprobar_solicitud(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    sol.estado = 'aprobada'
    sol.aprobado_por = request.user
    sol.fecha_aprobacion = timezone.now()
    sol.save()
    messages.success(request, f'Solicitud {sol.id_corto()} aprobada.')
    return redirect('materiales_panel')


@login_required
@require_POST
def editar_observacion(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    sol.observaciones = request.POST.get('observaciones', '')
    sol.save()
    messages.success(request, 'Observación guardada.')
    return redirect('detalle_solicitud', pk=pk)


@login_required
def exportar_excel_view(request):
    qs = Solicitud.objects.select_related('cuadrilla').prefetch_related('items__material')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    if estado: qs = qs.filter(estado=estado)
    if desde: qs = qs.filter(fecha_solicitud__gte=desde)
    if hasta: qs = qs.filter(fecha_solicitud__lte=hasta)
    out = exportar_solicitudes_excel(qs)
    resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="solicitudes.xlsx"'
    return resp


# ── PANEL ALMACÉN ─────────────────────────────────────────────────────────────

@login_required
def almacen_panel(request):
    # Solo almacen o superusuario
    if not request.user.is_superuser and request.user.username != 'almacen':
        messages.error(request, 'No tiene acceso al panel de almacén.')
        return redirect('materiales_panel')
    
    pendientes = Solicitud.objects.filter(
        estado__in=['aprobada', 'en_entrega']
    ).select_related('cuadrilla').prefetch_related('items__material').order_by('cuadrilla__movil', 'fecha_solicitud')
    
    historial = Solicitud.objects.filter(
    estado__in=['entregada', 'parcial', 'cerrada'],
    oculto_almacen=False
    ).select_related('cuadrilla').prefetch_related('items__material').order_by('-actualizado')[:50]

    return render(request, 'materiales/almacen_panel.html', {
        'solicitudes': pendientes,
        'historial': historial,
    })

@login_required
def almacen_entregar(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    if request.method == 'POST':
        accion = request.POST.get('accion', 'alistar')

        if accion == 'cerrar':
            sol.estado = 'cerrada'
            sol.save()
            messages.warning(request, 'Solicitud cerrada.')
            return redirect('almacen_panel')

        # Guardar cantidades entregadas
        for item in sol.items.all():
            cant = request.POST.get(f'cantidad_{item.pk}', '0')
            try:
                item.cantidad_entregada = float(cant)
            except ValueError:
                item.cantidad_entregada = 0
            item.estado_entrega = 'alistado'
            item.save()

        sol.estado = 'alistado'
        sol.save()
        messages.success(request, 'Solicitud marcada como alistada. El técnico debe confirmar la recepción.')
        return redirect('almacen_panel')

    return render(request, 'materiales/almacen_entregar.html', {'sol': sol})
def tecnico_confirmar_recepcion(request, pk):
    cuadrilla = _get_cuadrilla(request)
    if not cuadrilla:
        return redirect('tecnico_login')
    sol = get_object_or_404(Solicitud, pk=pk, cuadrilla=cuadrilla)
    if sol.estado != 'alistado':
        messages.error(request, 'Esta solicitud no está lista para confirmar.')
        return redirect('tecnico_solicitudes')
    if request.method == 'POST':
        for item in sol.items.all():
            item.estado_entrega = 'entregado'
            item.save()
        sol.estado = 'entregada'
        sol.save()
        messages.success(request, '¡Recepción confirmada! Gracias.')
        return redirect('tecnico_solicitudes')
    return render(request, 'materiales/tecnico_confirmar.html', {'sol': sol})

# ── API BÚSQUEDA MATERIALES ───────────────────────────────────────────────────

@require_GET
def api_buscar_material(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    materiales = Material.objects.filter(
        Q(descripcion__icontains=q) | Q(codigo__icontains=q),
        activo=True
    )[:15]
    return JsonResponse({'results': [
        {'id': m.pk, 'codigo': m.codigo, 'descripcion': m.descripcion, 'unidad': m.unidad}
        for m in materiales
    ]})


# ── PANEL TÉCNICO (acceso por código) ────────────────────────────────────────

def tecnico_login(request):
    error = ''
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        try:
            cuadrilla = Cuadrilla.objects.get(codigo=codigo, activo=True)
            request.session['cuadrilla_id'] = cuadrilla.pk
            request.session['cuadrilla_nombre'] = cuadrilla.nombre
            request.session['cuadrilla_movil'] = cuadrilla.movil
            return redirect('tecnico_solicitudes')
        except Cuadrilla.DoesNotExist:
            error = 'Código no encontrado. Intente nuevamente.'
    return render(request, 'materiales/tecnico_login.html', {'error': error})


def tecnico_logout(request):
    request.session.pop('cuadrilla_id', None)
    return redirect('tecnico_login')


def _get_cuadrilla(request):
    cid = request.session.get('cuadrilla_id')
    if not cid:
        return None
    try:
        return Cuadrilla.objects.get(pk=cid)
    except Cuadrilla.DoesNotExist:
        return None


def tecnico_solicitudes(request):
    cuadrilla = _get_cuadrilla(request)
    if not cuadrilla:
        return redirect('tecnico_login')
    pendientes = Solicitud.objects.filter(
        cuadrilla=cuadrilla
    ).exclude(estado='cerrada').prefetch_related('items__material')
    return render(request, 'materiales/tecnico_solicitudes.html', {
        'cuadrilla': cuadrilla, 'solicitudes': pendientes,
    })


def tecnico_historial(request):
    cuadrilla = _get_cuadrilla(request)
    if not cuadrilla:
        return redirect('tecnico_login')
    todas = Solicitud.objects.filter(cuadrilla=cuadrilla).prefetch_related('items__material')
    return render(request, 'materiales/tecnico_historial.html', {
        'cuadrilla': cuadrilla, 'solicitudes': todas,
    })


def tecnico_nueva_solicitud(request):
    cuadrilla = _get_cuadrilla(request)
    if not cuadrilla:
        return redirect('tecnico_login')
    if request.method == 'POST':
        items_json = request.POST.get('items_json', '[]')
        try:
            items = json.loads(items_json)
        except json.JSONDecodeError:
            items = []
        if not items:
            messages.error(request, 'Debe agregar al menos un material.')
            return render(request, 'materiales/tecnico_nueva_solicitud.html', {'cuadrilla': cuadrilla})

        sol = Solicitud.objects.create(cuadrilla=cuadrilla)
        for item in items:
            try:
                mat = Material.objects.get(pk=item['material_id'])
                ItemSolicitud.objects.create(
                    solicitud=sol,
                    material=mat,
                    cantidad_solicitada=float(item.get('cantidad', 1)),
                    acobro=item.get('acobro', 'sin_cobro'),
                    serie=item.get('serie', ''),
                )
            except (Material.DoesNotExist, KeyError, ValueError):
                continue
        messages.success(request, f'Solicitud {sol.id_corto()} creada exitosamente.')
        return redirect('tecnico_solicitudes')
    return render(request, 'materiales/tecnico_nueva_solicitud.html', {'cuadrilla': cuadrilla})


def tecnico_editar_solicitud(request, pk):
    cuadrilla = _get_cuadrilla(request)
    if not cuadrilla:
        return redirect('tecnico_login')
    sol = get_object_or_404(Solicitud, pk=pk, cuadrilla=cuadrilla)
    if not sol.puede_editar():
        messages.error(request, 'Esta solicitud ya no puede editarse.')
        return redirect('tecnico_solicitudes')
    if request.method == 'POST':
        items_json = request.POST.get('items_json', '[]')
        try:
            items = json.loads(items_json)
        except json.JSONDecodeError:
            items = []
        sol.items.all().delete()
        for item in items:
            try:
                mat = Material.objects.get(pk=item['material_id'])
                ItemSolicitud.objects.create(
                    solicitud=sol,
                    material=mat,
                    cantidad_solicitada=float(item.get('cantidad', 1)),
                    acobro=item.get('acobro', 'sin_cobro'),
                    serie=item.get('serie', ''),
                )
            except (Material.DoesNotExist, KeyError, ValueError):
                continue
        messages.success(request, 'Solicitud actualizada.')
        return redirect('tecnico_solicitudes')
    return render(request, 'materiales/tecnico_nueva_solicitud.html', {
        'cuadrilla': cuadrilla, 'sol': sol,
        'items_existentes': json.dumps([{
            'material_id': i.material.pk,
            'codigo': i.material.codigo,
            'descripcion': i.material.descripcion,
            'unidad': i.material.unidad,
            'cantidad': float(i.cantidad_solicitada),
            'acobro': i.acobro,
            'serie': i.serie,
        } for i in sol.items.select_related('material')])
    })


# ── ADMIN: gestión cuadrillas y materiales ───────────────────────────────────

@login_required
def gestion_cuadrillas(request):
    cuadrillas = Cuadrilla.objects.all()
    return render(request, 'materiales/cuadrillas.html', {'cuadrillas': cuadrillas})


@login_required
@require_POST
def crear_cuadrilla(request):
    codigo = request.POST.get('codigo', '').strip()
    nombre = request.POST.get('nombre', '').strip()
    movil = request.POST.get('movil', '').strip()
    supervisor = request.POST.get('Supervisor', '').strip()
    if codigo and nombre and movil:
        Cuadrilla.objects.update_or_create(
            codigo=codigo,
            defaults={
                'nombre': nombre,
                'movil': movil,
                'supervisor': supervisor,
                'activo': True
            }
        )
        messages.success(request, f'Cuadrilla {movil} guardada.')
    else:
        messages.error(request, 'Complete todos los campos.')
    return redirect('gestion_cuadrillas')


@login_required
@require_POST
def importar_materiales_view(request):
    import os
    ruta = os.path.join('materiales', 'lista_de_materiales_hesego.xlsx')
    stats = importar_materiales(ruta)
    messages.success(request, f"Materiales importados: {stats['creados']} nuevos, {stats['actualizados']} actualizados.")
    return redirect('panel_materiales_admin')

@login_required
def exportar_plantilla_cuadrilla(request):
    """Exporta una solicitud específica en el formato de Plantilla_Almacen."""
    from copy import copy
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.conf import settings
    import os

    cuadrilla_id = request.GET.get('cuadrilla', '')
    fecha = request.GET.get('fecha', '')

    if not cuadrilla_id or not fecha:
        messages.error(request, 'Seleccione cuadrilla y fecha.')
        return redirect('lista_solicitudes')

    solicitudes = Solicitud.objects.filter(
        cuadrilla_id=cuadrilla_id,
        fecha_solicitud=fecha,
    ).prefetch_related('items__material')

    if not solicitudes.exists():
        messages.error(request, 'No hay solicitudes para esa cuadrilla y fecha.')
        return redirect('lista_solicitudes')

    # Cargar plantilla base
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'materiales', 'Plantilla_Almacen_-_copia.xlsx')
    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb['Plantilla']

    sol = solicitudes.first()

    # Llenar encabezado
    ws['A6'] = fecha
    ws['A10'] = sol.cuadrilla.nombre
    ws['E10'] = sol.cuadrilla.movil

    # Fila donde empiezan los items (fila 23 según la plantilla)
    fila_inicio = 23
    for idx, item in enumerate(sol.items.select_related('material'), start=0):
        fila = fila_inicio + idx
        ws.cell(row=fila, column=1, value=idx + 1)           # POS
        ws.cell(row=fila, column=2, value=item.material.codigo)       # Código
        ws.cell(row=fila, column=3, value=item.material.descripcion)  # Descripción
        ws.cell(row=fila, column=4, value=item.material.unidad)       # UM
        ws.cell(row=fila, column=5, value=float(item.cantidad_solicitada))  # Cant. Solicitada
        ws.cell(row=fila, column=6, value=float(item.cantidad_entregada))   # Cant. Entregada
        ws.cell(row=fila, column=10, value=item.serie or '')          # Series

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    nombre_archivo = f"solicitud_{sol.cuadrilla.movil}_{fecha}.xlsx"
    resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return resp

def es_almacen(user):
    return user.username == 'almacen' or user.is_superuser

@login_required
def gestion_materiales(request):
    if not es_almacen(request.user):
        messages.error(request, 'Sin acceso.')
        return redirect('materiales_panel')
    materiales = Material.objects.all().order_by('descripcion')
    return render(request, 'materiales/gestion_materiales.html', {'materiales': materiales})


@login_required
@require_POST
def guardar_material(request, pk=None):
    if not es_almacen(request.user):
        return JsonResponse({'ok': False})
    codigo = request.POST.get('codigo', '').strip()
    descripcion = request.POST.get('descripcion', '').strip()
    unidad = request.POST.get('unidad', '').strip()
    if not codigo or not descripcion:
        messages.error(request, 'Código y descripción son obligatorios.')
        return redirect('gestion_materiales')
    if pk:
        mat = get_object_or_404(Material, pk=pk)
        mat.codigo = codigo
        mat.descripcion = descripcion
        mat.unidad = unidad
        mat.save()
        messages.success(request, 'Material actualizado.')
    else:
        Material.objects.create(codigo=codigo, descripcion=descripcion, unidad=unidad, activo=True)
        messages.success(request, 'Material agregado.')
    return redirect('gestion_materiales')


@login_required
@require_POST
def eliminar_material(request, pk):
    if not es_almacen(request.user):
        return JsonResponse({'ok': False})
    mat = get_object_or_404(Material, pk=pk)
    mat.delete()
    messages.success(request, 'Material eliminado.')
    return redirect('gestion_materiales')

@login_required
@require_POST
def editar_cuadrilla(request, pk):
    cuadrilla = get_object_or_404(Cuadrilla, pk=pk)
    cuadrilla.codigo = request.POST.get('codigo', '').strip()
    cuadrilla.nombre = request.POST.get('nombre', '').strip()
    cuadrilla.movil = request.POST.get('movil', '').strip()
    cuadrilla.supervisor = request.POST.get('Supervisor', '').strip()
    cuadrilla.activo = request.POST.get('activo') == 'on'
    cuadrilla.save()
    messages.success(request, f'Cuadrilla {cuadrilla.movil} actualizada.')
    return redirect('gestion_cuadrillas')


@login_required
@require_POST
def eliminar_cuadrilla(request, pk):
    cuadrilla = get_object_or_404(Cuadrilla, pk=pk)
    movil = cuadrilla.movil
    cuadrilla.delete()
    messages.success(request, f'Cuadrilla {movil} eliminada.')
    return redirect('gestion_cuadrillas')


@login_required
@require_POST
def ocultar_solicitud_almacen(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    sol.oculto_almacen = True
    sol.save()
    return JsonResponse({'ok': True})


@login_required
def descargar_pdf_solicitud(request, pk):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from io import BytesIO

    sol = get_object_or_404(Solicitud, pk=pk)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    titulo = ParagraphStyle('titulo', fontSize=13, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)
    subtitulo = ParagraphStyle('sub', fontSize=10, fontName='Helvetica-Bold', alignment=1, spaceAfter=6)
    normal = ParagraphStyle('normal', fontSize=9, fontName='Helvetica', spaceAfter=3)
    azul = colors.HexColor('#1a3a5c')

    elementos = []

    # Encabezado empresa
    elementos.append(Paragraph('HESEGO', titulo))
    elementos.append(Paragraph('NIT: 900502031-8', subtitulo))
    elementos.append(Paragraph('Cúcuta', subtitulo))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph('SALIDA DE ALMACÉN', titulo))
    elementos.append(Spacer(1, 0.4*cm))

    # Info solicitud
    info_data = [
        ['Fecha Movimiento:', sol.fecha_solicitud.strftime('%d/%m/%Y'),
         'Cuadrilla:', sol.cuadrilla.movil],
        ['Concepto:', 'ASIGNACIÓN DE MATERIALES', '', ''],
        ['Bodega:', f'{sol.cuadrilla.supervisor or "—"} - {sol.cuadrilla.nombre}',
         'Técnico:', sol.cuadrilla.nombre],
    ]
    tabla_info = Table(info_data, colWidths=[3.5*cm, 6*cm, 3*cm, 5*cm])
    tabla_info.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_info)
    elementos.append(Spacer(1, 0.5*cm))

    # Tabla de materiales
    headers = ['POS', 'Código', 'Descripción del Material', 'UM', 'Cant. Solicitada', 'Cant. Entregada']
    data = [headers]
    for idx, item in enumerate(sol.items.select_related('material'), 1):
        data.append([
            str(idx),
            item.material.codigo,
            item.material.descripcion,
            item.material.unidad,
            str(item.cantidad_solicitada),
            str(item.cantidad_entregada),
        ])

    tabla = Table(data, colWidths=[1*cm, 3*cm, 8*cm, 1.5*cm, 2.5*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 1.5*cm))

    # Firmas
    firmas = [['Entregó', '', 'Recibió']]
    tabla_firmas = Table(firmas, colWidths=[6*cm, 5*cm, 6*cm])
    tabla_firmas.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ('LINEABOVE', (2, 0), (2, 0), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 30),
    ]))
    elementos.append(tabla_firmas)

    doc.build(elementos)
    buffer.seek(0)
    nombre = f"salida_almacen_{sol.cuadrilla.movil}_{sol.fecha_solicitud}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response

@login_required
@require_POST
def eliminar_solicitud(request, pk):
    sol = get_object_or_404(Solicitud, pk=pk)
    sol.delete()
    messages.success(request, 'Solicitud eliminada.')
    return redirect('materiales_panel')