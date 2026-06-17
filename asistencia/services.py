"""
Servicios de negocio: OTP, SMS, QR, geolocalización, importación Excel.
"""
import random
import string
import secrets
import math
from datetime import timedelta

import qrcode
from io import BytesIO
import base64

from django.utils import timezone
from django.conf import settings

from .models import Tecnico, CodigoQR, OTPRegistro


# ─── OTP ────────────────────────────────────────────────────────────────────

def generar_otp():
    """Genera un código OTP numérico de 6 dígitos."""
    return ''.join(random.choices(string.digits, k=6))


def crear_otp(tecnico: Tecnico) -> OTPRegistro:
    """Invalida OTPs anteriores del técnico y crea uno nuevo."""
    OTPRegistro.objects.filter(tecnico=tecnico, usado=False).update(usado=True)
    expira = timezone.now() + timedelta(minutes=settings.OTP_EXPIRY_MINUTES)
    otp = OTPRegistro.objects.create(
        tecnico=tecnico,
        codigo=generar_otp(),
        expira=expira,
    )
    return otp


def verificar_otp(tecnico: Tecnico, codigo: str) -> tuple[bool, str]:
    """
    Verifica el OTP ingresado por el técnico.
    Retorna (valido: bool, mensaje: str).
    """
    try:
        otp = OTPRegistro.objects.filter(
            tecnico=tecnico, codigo=codigo, usado=False
        ).latest('creado')
    except OTPRegistro.DoesNotExist:
        return False, "Código OTP incorrecto."

    if not otp.esta_vigente():
        return False, "El código OTP ha expirado. Solicite uno nuevo."

    otp.usado = True
    otp.save()
    return True, "OTP verificado correctamente."


# ─── SMS ─────────────────────────────────────────────────────────────────────

def enviar_sms_otp(celular: str, codigo: str) -> bool:
    """
    Envía el OTP por SMS.
    En desarrollo usa el backend 'console' (imprime en logs).
    En producción configura SMS_BACKEND=twilio y las variables de Twilio.
    """
    mensaje = f"Su código de asistencia es: {codigo}. Válido por {settings.OTP_EXPIRY_MINUTES} minutos."

    if settings.SMS_BACKEND == 'twilio':
        try:
            from twilio.rest import Client
            client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
            client.messages.create(
                body=mensaje,
                from_=settings.TWILIO_PHONE_NUMBER,
                to=f'+57{celular}'
            )
            return True
        except Exception as e:
            print(f"[SMS ERROR] {e}")
            return False
    else:
        # Backend consola: en desarrollo el OTP se muestra en pantalla
        print(f"[SMS CONSOLA] Para {celular}: {mensaje}")
        return True


# ─── QR ──────────────────────────────────────────────────────────────────────

def generar_token_qr() -> str:
    """Genera un token seguro y único para el QR."""
    return secrets.token_urlsafe(32)


def crear_qr(request, dias_vigencia: int = 30) -> CodigoQR:
    """
    Desactiva todos los QR anteriores y crea uno nuevo.
    Genera y guarda la imagen del QR en media/qrcodes/.
    """
    CodigoQR.objects.filter(activo=True).update(activo=False)

    token = generar_token_qr()
    fecha_vencimiento = timezone.now() + timedelta(days=dias_vigencia)

    qr_obj = CodigoQR.objects.create(
        token=token,
        activo=True,
        fecha_vencimiento=fecha_vencimiento,
        creado_por=request.user if request.user.is_authenticated else None,
    )

    # Generar imagen QR
    url_registro = request.build_absolute_uri(f'/registrar/{token}/')
    _guardar_imagen_qr(token, url_registro)

    return qr_obj


def _guardar_imagen_qr(token: str, url: str):
    """Genera y guarda la imagen PNG del QR."""
    import os
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    ruta = os.path.join(settings.MEDIA_ROOT, 'qrcodes', f'qr_{token[:16]}.png')
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    img.save(ruta)


def obtener_qr_imagen_base64(token: str) -> str:
    """Retorna la imagen QR como base64 para mostrar en HTML."""
    import os
    ruta = os.path.join(settings.MEDIA_ROOT, 'qrcodes', f'qr_{token[:16]}.png')
    if os.path.exists(ruta):
        with open(ruta, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    return ''


# ─── GEOLOCALIZACIÓN ─────────────────────────────────────────────────────────

def calcular_distancia(lat1, lon1, lat2, lon2) -> float:
    """
    Calcula distancia en metros entre dos coordenadas (fórmula Haversine).
    """
    R = 6371000  # Radio de la Tierra en metros
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def validar_ubicacion(lat: float, lng: float) -> tuple[bool, int]:
    """
    Verifica si las coordenadas están dentro del radio autorizado.
    Retorna (dentro: bool, distancia_metros: int).
    """
    distancia = calcular_distancia(lat, lng, settings.GEO_LAT, settings.GEO_LNG)
    return distancia <= settings.GEO_RADIO_METROS, int(distancia)


# ─── IMPORTACIÓN EXCEL ───────────────────────────────────────────────────────

def importar_tecnicos_excel(ruta_excel: str) -> dict:
    """
    Importa/actualiza técnicos desde un archivo Excel.
    Columnas esperadas: CEDULA, NOMBRE, CELULAR.
    Retorna estadísticas del proceso.
    """
    import openpyxl
    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb.active

    creados, actualizados, errores = 0, 0, 0

    # Saltar encabezado
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        cedula, nombre, celular = row[0], row[1], row[2]
        if not cedula or not nombre:
            continue
        try:
            obj, created = Tecnico.objects.update_or_create(
                cedula=str(cedula).strip(),
                defaults={
                    'nombre': str(nombre).strip(),
                    'celular': str(int(celular)).strip() if celular else '',
                    'activo': True,
                }
            )
            if created:
                creados += 1
            else:
                actualizados += 1
        except Exception as e:
            print(f"[IMPORT ERROR] Fila {cedula}: {e}")
            errores += 1

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}


# ─── EXPORTACIÓN EXCEL ───────────────────────────────────────────────────────

def exportar_asistencia_excel(queryset) -> BytesIO:
    """
    Exporta un queryset de RegistroAsistencia a un archivo Excel en memoria.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    # Encabezados
    headers = [
        'Cédula', 'Nombre', 'Celular', 'Fecha', 'Hora Registro',
        'Estado', 'Tarde', 'Latitud', 'Longitud', 'Distancia (m)',
        'OTP Utilizado', 'IP Dispositivo'
    ]
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for row_idx, reg in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=reg.tecnico.cedula)
        ws.cell(row=row_idx, column=2, value=reg.tecnico.nombre)
        ws.cell(row=row_idx, column=3, value=reg.tecnico.celular)
        ws.cell(row=row_idx, column=4, value=reg.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=5, value=reg.hora_registro.strftime('%H:%M:%S'))
        ws.cell(row=row_idx, column=6, value=reg.get_estado_display())
        ws.cell(row=row_idx, column=7, value='Sí' if reg.tarde else 'No')
        ws.cell(row=row_idx, column=8, value=float(reg.latitud) if reg.latitud else '')
        ws.cell(row=row_idx, column=9, value=float(reg.longitud) if reg.longitud else '')
        ws.cell(row=row_idx, column=10, value=reg.distancia_metros or '')
        ws.cell(row=row_idx, column=11, value=reg.otp_utilizado)
        ws.cell(row=row_idx, column=12, value=reg.ip_dispositivo or '')

        # Resaltar tardanzas
        if reg.tarde:
            for col in range(1, 13):
                ws.cell(row=row_idx, column=col).fill = PatternFill('solid', start_color='FFE0B2')

    # Ajustar anchos
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
